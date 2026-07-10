from collections import Counter, defaultdict
from datetime import datetime, time
from io import BytesIO
from unicodedata import normalize

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from conexao import db
from model import Recurso, Reserva, Setor, TipoRecurso, Usuario
from services.pdf import gerar_pdf, gerar_pdf_historico


main = Blueprint("main", __name__)


def role_atual():
    return (session.get("role") or "").strip().lower()


def usuario_tecnico():
    return role_atual() == "tecnico"


def usuario_rh():
    return role_atual() == "rh"


def pode_ver_relatorios():
    return role_atual() in ["rh", "diretoria", "tecnico"]


def exigir_tecnico():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not usuario_tecnico():
        return "Acesso negado", 403

    return None


def usuario_logado():
    if "usuario" not in session:
        return None

    return Usuario.query.filter_by(usuario=session["usuario"]).first()


def consulta_reservas_filtrada():
    data_inicio = request.args.get("dataInicio")
    data_fim = request.args.get("dataFim")
    setor = request.args.get("setor")

    consulta = Reserva.query

    if data_inicio:
        consulta = consulta.filter(
            Reserva.data_reserva >= datetime.strptime(data_inicio, "%Y-%m-%d").date()
        )

    if data_fim:
        consulta = consulta.filter(
            Reserva.data_reserva <= datetime.strptime(data_fim, "%Y-%m-%d").date()
        )

    if setor:
        consulta = consulta.filter(Reserva.setor.contains(setor))

    return consulta, {
        "dataInicio": data_inicio,
        "dataFim": data_fim,
        "setor": setor,
    }


def montar_stats(registros):
    total = len(registros)
    contagem_setores = Counter(registro.setor for registro in registros if registro.setor)
    setor_top = contagem_setores.most_common(1)[0][0] if contagem_setores else "-"

    datas_unicas = set(registro.data_reserva for registro in registros)
    media_diaria = round(total / len(datas_unicas), 1) if datas_unicas else 0

    contagem_responsaveis = Counter(
        registro.responsavel for registro in registros if registro.responsavel
    )
    responsavel_top = (
        contagem_responsaveis.most_common(1)[0][0] if contagem_responsaveis else "-"
    )
    contagem_recursos = Counter(
        registro.recurso.nome for registro in registros if registro.recurso
    )
    recurso_top = contagem_recursos.most_common(1)[0][0] if contagem_recursos else "-"
    status_contagem = Counter(registro.status_calculado for registro in registros)
    devolvidos = status_contagem.get("devolvido", 0)
    cancelados = status_contagem.get("cancelado", 0)
    pendentes = max(total - devolvidos - cancelados, 0)
    taxa_devolucao = round((devolvidos / total) * 100, 1) if total else 0
    dias_com_uso = len(datas_unicas)

    return {
        "total": total,
        "setorTop": setor_top,
        "mediaDiaria": media_diaria,
        "requerenteTop": responsavel_top,
        "recursoTop": recurso_top,
        "emUso": status_contagem.get("usando", 0),
        "atrasados": status_contagem.get("atrasado", 0),
        "devolvidos": devolvidos,
        "pendentes": pendentes,
        "cancelados": cancelados,
        "viagens": status_contagem.get("viagem", 0),
        "reservados": status_contagem.get("reservado", 0),
        "taxaDevolucao": taxa_devolucao,
        "diasComUso": dias_com_uso,
    }


def montar_dados_relatorio(registros):
    contagem_setores = Counter(registro.setor for registro in registros if registro.setor)
    contagem_recursos = Counter(
        registro.recurso.nome for registro in registros if registro.recurso
    )
    contagem_responsaveis = Counter(registro.responsavel for registro in registros if registro.responsavel)
    contagem_status = Counter(registro.status_label for registro in registros)
    registro_por_data = defaultdict(int)
    registro_por_hora = defaultdict(int)

    for registro in registros:
        registro_por_data[registro.data_reserva.strftime("%d/%m")] += 1
        registro_por_hora[registro.hora_inicio.strftime("%H:00")] += 1

    periodo_ordenado = sorted(
        registro_por_data.items(),
        key=lambda item: datetime.strptime(item[0], "%d/%m"),
    )
    horas_ordenadas = sorted(registro_por_hora.items())

    return {
        "setor": {
            "labels": list(contagem_setores.keys()),
            "valores": list(contagem_setores.values()),
        },
        "periodo": {
            "labels": [label for label, _ in periodo_ordenado],
            "valores": [valor for _, valor in periodo_ordenado],
        },
        "recurso": {
            "labels": list(contagem_recursos.keys()),
            "valores": list(contagem_recursos.values()),
        },
        "status": {
            "labels": list(contagem_status.keys()),
            "valores": list(contagem_status.values()),
        },
        "hora": {
            "labels": [label for label, _ in horas_ordenadas],
            "valores": [valor for _, valor in horas_ordenadas],
        },
        "responsavel": {
            "labels": list(contagem_responsaveis.keys()),
            "valores": list(contagem_responsaveis.values()),
        },
        "rankings": {
            "recursos": contagem_recursos.most_common(5),
            "setores": contagem_setores.most_common(5),
            "responsaveis": contagem_responsaveis.most_common(5),
        },
    }


def reservas_ativas_do_recurso(recurso_id):
    return (
        Reserva.query.filter_by(recurso_id=recurso_id)
        .filter(Reserva.status.notin_(["devolvido", "cancelado"]))
        .order_by(Reserva.data_reserva.asc(), Reserva.hora_inicio.asc())
        .all()
    )


def montar_agenda_ocupada(recurso_id):
    if not recurso_id:
        return []

    agenda = defaultdict(list)

    for reserva in reservas_ativas_do_recurso(recurso_id):
        hora_fim = reserva.hora_fim.strftime("%H:%M") if reserva.hora_fim else "Viagem"
        agenda[reserva.data_reserva].append(
            {
                "inicio": reserva.hora_inicio.strftime("%H:%M"),
                "fim": hora_fim,
                "responsavel": reserva.responsavel or "Sem responsavel",
                "status": reserva.status_label,
            }
        )

    return [
        {
            "data": data,
            "data_formatada": data.strftime("%d/%m/%Y"),
            "horarios": horarios,
        }
        for data, horarios in agenda.items()
    ]


def horarios_conflitam(inicio_a, fim_a, viagem_a, inicio_b, fim_b, viagem_b):
    if viagem_a or viagem_b:
        return True

    fim_a = fim_a or time(18, 0)
    fim_b = fim_b or time(18, 0)

    return inicio_a < fim_b and fim_a > inicio_b


def existe_conflito_reserva(recurso_id, data_reserva, hora_inicio, hora_fim, viagem):
    reservas_do_dia = (
        Reserva.query.filter_by(recurso_id=recurso_id, data_reserva=data_reserva)
        .filter(Reserva.status.notin_(["devolvido", "cancelado"]))
        .all()
    )

    for reserva in reservas_do_dia:
        if horarios_conflitam(
            hora_inicio,
            hora_fim,
            viagem,
            reserva.hora_inicio,
            reserva.hora_fim,
            reserva.viagem,
        ):
            return reserva

    return None


def recurso_eh_veiculo(recurso):
    if not recurso or not recurso.tipo_recurso:
        return False

    tipo = normalizar_texto(recurso.tipo_recurso.nome)
    return tipo.startswith("veiculo")


def normalizar_texto(texto):
    return normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii").strip().lower()


def recurso_controlado_pelo_rh(recurso):
    if not recurso or not recurso.tipo_recurso:
        return False

    tipo = normalizar_texto(recurso.tipo_recurso.nome)
    return tipo.startswith("sala") or tipo.startswith("veiculo")


def recurso_disponivel_para_reserva(recurso):
    if not recurso or not recurso.ativo:
        return False

    return normalizar_texto(recurso.status) == "disponivel"


def consulta_recursos_disponiveis():
    return (
        Recurso.query.filter_by(ativo=True)
        .filter(db.func.lower(db.func.trim(Recurso.status)) == "disponivel")
        .order_by(Recurso.nome.asc())
    )


def consulta_setores_ativos():
    return Setor.query.filter_by(ativo=True).order_by(Setor.nome.asc())


def prioridade_reserva_atrasada():
    agora = datetime.now()
    return db.case(
        (
            (Reserva.status.notin_(["devolvido", "cancelado"]))
            & (Reserva.hora_fim.isnot(None))
            & (
                (Reserva.data_reserva < agora.date())
                | (
                    (Reserva.data_reserva == agora.date())
                    & (Reserva.hora_fim < agora.time())
                )
            ),
            0,
        ),
        else_=1,
    )


def ordenar_historico(consulta):
    return consulta.order_by(
        prioridade_reserva_atrasada().asc(),
        Reserva.data_reserva.desc(),
        Reserva.hora_inicio.desc(),
    )


@main.route("/")
def index():
    return redirect(url_for("main.home"))


@main.route("/home")
def home():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    pagina = request.args.get("page", 1, type=int)

    consulta_agendamentos = Reserva.query
    consulta_abertos = Reserva.query

    if not usuario_tecnico():
        consulta_agendamentos = consulta_agendamentos.filter_by(usuario_id=session.get("usuario_id"))
        consulta_abertos = consulta_abertos.filter_by(usuario_id=session.get("usuario_id"))

    agendamentos = (
        ordenar_historico(consulta_agendamentos)
        .paginate(
        page=pagina,
        per_page=10,
        error_out=False,
        )
    )

    emprestimos_abertos = (
        consulta_abertos.filter(Reserva.status.notin_(["devolvido", "cancelado"]))
        .order_by(Reserva.data_reserva.desc())
        .all()
    )

    usuarios = Usuario.query.order_by(Usuario.usuario.asc()).all() if usuario_tecnico() else []
    setores = consulta_setores_ativos().all() if usuario_tecnico() else []

    return render_template(
        "home.html",
        usuario=session["usuario"],
        agendamentos=agendamentos,
        emprestimos_abertos=emprestimos_abertos,
        usuarios=usuarios,
        setores=setores,
    )


@main.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("main.login_page"))


@main.route("/login", methods=["GET"])
def login_page():
    return render_template("login.html")


@main.route("/login", methods=["POST"])
def login():
    usuario_form = request.form["usuario"]
    senha_form = request.form["senha"]

    user = Usuario.query.filter_by(usuario=usuario_form).first()

    if not user:
        flash("Usuario nao encontrado", "erro")
        return redirect(url_for("main.login_page"))

    if user.senha != senha_form:
        flash("Senha incorreta", "erro")
        return redirect(url_for("main.login_page"))

    session["usuario_id"] = user.id
    session["usuario"] = user.usuario
    session["nome"] = user.usuario
    session["role"] = user.role

    return redirect(url_for("main.home"))


@main.route("/cadastro", methods=["GET"])
def cadastro_page():
    return render_template("cadastro.html")


@main.route("/cadastro", methods=["POST"])
def cadastro():
    novo_usuario = Usuario(
        usuario=request.form["usuario"],
        senha=request.form["senha"],
    )

    db.session.add(novo_usuario)
    db.session.commit()

    session["usuario_id"] = novo_usuario.id
    session["usuario"] = novo_usuario.usuario
    session["nome"] = novo_usuario.usuario
    session["role"] = novo_usuario.role

    return redirect(url_for("main.home"))


@main.route("/relatorios")
def relatorios():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not pode_ver_relatorios():
        return "Acesso negado", 403

    consulta, filtros = consulta_reservas_filtrada()
    registros = consulta.all()
    stats = montar_stats(registros)
    dados = montar_dados_relatorio(registros)

    return render_template(
        "relatorios.html",
        filtros=filtros,
        stats=stats,
        dados_setor=dados["setor"],
        dados_periodo=dados["periodo"],
        dados_recurso=dados["recurso"],
        dados_status=dados["status"],
        dados_hora=dados["hora"],
        dados_responsavel=dados["responsavel"],
        rankings=dados["rankings"],
    )


@main.route("/exportar/excel")
def exportar_excel():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not pode_ver_relatorios():
        return "Acesso negado", 403

    consulta, _ = consulta_reservas_filtrada()
    registros = consulta.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Base de dados"

    ws.append(["Data", "Recurso", "Responsavel", "Setor", "Local/Motivo", "Status", "Hora inicial", "Hora final", "Observacao"])

    for registro in registros:
        ws.append(
            [
                registro.data_reserva.strftime("%d/%m/%Y"),
                registro.recurso.nome if registro.recurso else "",
                registro.responsavel,
                registro.setor,
                registro.motivo,
                registro.status_label,
                registro.hora_inicio.strftime("%H:%M"),
                registro.hora_fim.strftime("%H:%M") if registro.hora_fim else "Viagem",
                registro.observacao,
            ]
        )

    header_fill = PatternFill("solid", fgColor="00995C")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E8DE"),
        right=Side(style="thin", color="D9E8DE"),
        top=Side(style="thin", color="D9E8DE"),
        bottom=Side(style="thin", color="D9E8DE"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 3, 12), 34)

    resumo = wb.create_sheet("Resumo")
    stats = montar_stats(registros)
    dados = montar_dados_relatorio(registros)
    resumo.append(["Indicador", "Valor"])
    resumo.append(["Total de registros", stats["total"]])
    resumo.append(["Dias com uso", stats["diasComUso"]])
    resumo.append(["Media diaria", stats["mediaDiaria"]])
    resumo.append(["Recurso mais usado", stats["recursoTop"]])
    resumo.append(["Setor mais ativo", stats["setorTop"]])
    resumo.append(["Responsavel mais frequente", stats["requerenteTop"]])
    resumo.append(["Taxa de devolucao", f"{stats['taxaDevolucao']}%"])
    resumo.append(["Pendentes", stats["pendentes"]])
    resumo.append(["Em uso", stats["emUso"]])
    resumo.append(["Atrasados", stats["atrasados"]])
    resumo.append(["Devolvidos", stats["devolvidos"]])
    resumo.append(["Viagens", stats["viagens"]])

    start_row = resumo.max_row + 2
    resumo.cell(start_row, 1, "Distribuicao por status")
    resumo.cell(start_row, 1).font = Font(bold=True, color="00995C")
    for index, (label, valor) in enumerate(zip(dados["status"]["labels"], dados["status"]["valores"]), start=start_row + 1):
        resumo.cell(index, 1, label)
        resumo.cell(index, 2, valor)

    for cell in resumo[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in resumo.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    resumo.column_dimensions["A"].width = 30
    resumo.column_dimensions["B"].width = 18
    resumo.sheet_view.showGridLines = False

    rankings = wb.create_sheet("Rankings")
    linha = 1
    for titulo, itens in [
        ("Top recursos", dados["rankings"]["recursos"]),
        ("Top setores", dados["rankings"]["setores"]),
        ("Top responsaveis", dados["rankings"]["responsaveis"]),
    ]:
        rankings.cell(linha, 1, titulo)
        rankings.cell(linha, 1).font = Font(bold=True, color="00995C", size=13)
        linha += 1
        rankings.append(["Nome", "Reservas"])
        for cell in rankings[linha]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        linha += 1
        for nome, valor in itens:
            rankings.cell(linha, 1, nome)
            rankings.cell(linha, 2, valor)
            rankings.cell(linha, 1).border = thin_border
            rankings.cell(linha, 2).border = thin_border
            linha += 1
        linha += 2

    rankings.column_dimensions["A"].width = 34
    rankings.column_dimensions["B"].width = 14
    rankings.sheet_view.showGridLines = False

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main.route("/exportar/pdf")
def exportar_pdf():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not pode_ver_relatorios():
        return "Acesso negado", 403

    consulta, filtros = consulta_reservas_filtrada()
    registros = consulta.all()
    stats = montar_stats(registros)

    dados = montar_dados_relatorio(registros)

    return gerar_pdf(registros, filtros, stats, dados)


@main.route("/historico/exportar/excel")
def exportar_historico_excel():
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    registros = ordenar_historico(Reserva.query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historico"

    ws.append(["Data", "Hora", "Responsavel", "Recurso", "Setor", "Local", "Status", "Observacao"])

    for registro in registros:
        hora_fim = registro.hora_fim.strftime("%H:%M") if registro.hora_fim else ("Viagem" if registro.viagem else "-")
        ws.append(
            [
                registro.data_reserva.strftime("%d/%m/%Y"),
                f"{registro.hora_inicio.strftime('%H:%M')} - {hora_fim}",
                registro.responsavel or (registro.usuario.usuario if registro.usuario else "-"),
                registro.recurso.nome if registro.recurso else "-",
                registro.setor or "-",
                registro.motivo or "-",
                registro.status_label,
                registro.observacao or "-",
            ]
        )

    header_fill = PatternFill("solid", fgColor="00995C")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E8DE"),
        right=Side(style="thin", color="D9E8DE"),
        top=Side(style="thin", color="D9E8DE"),
        bottom=Side(style="thin", color="D9E8DE"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 3, 12), 38)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="historico_geral_reservas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main.route("/historico/exportar/pdf")
def exportar_historico_pdf():
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    registros = ordenar_historico(Reserva.query).all()
    return gerar_pdf_historico(registros)


@main.route("/reservas", methods=["POST"])
@main.route("/datashow", methods=["POST"])
def salvar_reserva():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    recurso_id = request.form["recurso_id"]
    recurso = Recurso.query.get_or_404(recurso_id)

    if not recurso_disponivel_para_reserva(recurso):
        flash("Este equipamento nao esta disponivel para reserva no momento.", "erro")
        return redirect(url_for("main.reserva"))

    pode_viagem = recurso_eh_veiculo(recurso)
    viagem = request.form.get("viagem") == "on" and pode_viagem
    hora_fim_form = request.form.get("hora_fim")
    hora_fim = None
    setor = Setor.query.filter_by(id=request.form["setor_id"], ativo=True).first_or_404()
    data_reserva = datetime.strptime(request.form["data_reserva"], "%Y-%m-%d").date()
    hora_inicio = datetime.strptime(request.form["hora_inicio"], "%H:%M").time()

    if hora_fim_form:
        hora_fim = datetime.strptime(hora_fim_form, "%H:%M").time()
    elif not viagem:
        hora_fim = time(18, 0)

    if hora_fim and hora_fim <= hora_inicio:
        flash("A hora final precisa ser maior que a hora inicial.", "erro")
        return redirect(url_for("main.reserva", recurso_id=recurso_id))

    conflito = existe_conflito_reserva(
        recurso_id=recurso_id,
        data_reserva=data_reserva,
        hora_inicio=hora_inicio,
        hora_fim=hora_fim,
        viagem=viagem,
    )

    if conflito:
        fim_conflito = conflito.hora_fim.strftime("%H:%M") if conflito.hora_fim else "Viagem"
        flash(
            f"Este recurso ja esta reservado em {conflito.data_reserva.strftime('%d/%m/%Y')} "
            f"das {conflito.hora_inicio.strftime('%H:%M')} ate {fim_conflito}.",
            "erro",
        )
        return redirect(url_for("main.reserva", recurso_id=recurso_id))

    registro = Reserva(
        recurso_id=recurso_id,
        usuario_id=session.get("usuario_id"),
        responsavel=session.get("nome") or session.get("usuario"),
        setor=setor.nome,
        motivo=request.form.get("motivo"),
        data_reserva=data_reserva,
        hora_inicio=hora_inicio,
        hora_fim=hora_fim,
        observacao=request.form.get("observacao"),
        viagem=viagem,
    )

    db.session.add(registro)
    db.session.commit()

    return redirect(url_for("main.home"))


@main.route("/registro/<int:id>/excluir", methods=["POST"])
def excluir_registro(id):
    if not usuario_tecnico():
        return "Acesso negado", 403

    registro = Reserva.query.get_or_404(id)

    db.session.delete(registro)
    db.session.commit()

    return redirect(url_for("main.home"))


@main.route("/registro/<int:id>/editar", methods=["GET", "POST"])
def editar_registro(id):
    if not usuario_tecnico():
        return "Acesso negado", 403

    return redirect(url_for("main.home"))


@main.route("/registro/<int:id>/devolver", methods=["POST"])
def devolver_item(id):
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    registro = Reserva.query.get_or_404(id)

    if registro.usuario_id != session.get("usuario_id") and not usuario_tecnico():
        return "Acesso negado", 403

    registro.status = "devolvido"
    db.session.commit()

    return redirect(url_for("main.home"))


@main.route("/equipamentos")
def equipamentos():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    recursos = (
        Recurso.query.filter_by(ativo=True)
        .order_by(Recurso.nome.asc())
        .all()
    )
    recursos_disponiveis = consulta_recursos_disponiveis().all()
    recursos_rh = [recurso for recurso in recursos if recurso_controlado_pelo_rh(recurso)]
    tipos_recursos = TipoRecurso.query.filter_by(ativo=True).order_by(TipoRecurso.nome.asc()).all()

    return render_template(
        "equipamentos.html",
        recursos=recursos,
        recursos_disponiveis=recursos_disponiveis,
        recursos_rh=recursos_rh,
        tipos_recursos=tipos_recursos,
    )


@main.route("/equipamentos/adicionar", methods=["POST"])
def adicionar_equipamento():
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    novo_recurso = Recurso(
        tipo_recurso_id=request.form["tipo_recurso_id"],
        nome=request.form["nome"],
        descricao=request.form.get("descricao"),
        status=request.form.get("status") or "disponivel",
        ativo=True,
    )

    db.session.add(novo_recurso)
    db.session.commit()

    return redirect(url_for("main.equipamentos"))


@main.route("/equipamentos/<int:id>/editar", methods=["POST"])
def editar_equipamento(id):
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    recurso = Recurso.query.get_or_404(id)
    recurso.tipo_recurso_id = request.form["tipo_recurso_id"]
    recurso.nome = request.form["nome"]
    recurso.descricao = request.form.get("descricao")
    recurso.status = request.form.get("status") or "disponivel"

    db.session.commit()

    return redirect(url_for("main.equipamentos"))


@main.route("/equipamentos/<int:id>/status", methods=["POST"])
def alterar_status_equipamento(id):
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    recurso = Recurso.query.get_or_404(id)

    if not usuario_tecnico() and not (usuario_rh() and recurso_controlado_pelo_rh(recurso)):
        return "Acesso negado", 403

    status = request.form.get("status") or "disponivel"
    if status not in ["disponivel", "manutencao", "indisponivel"]:
        status = "disponivel"

    recurso.status = status
    db.session.commit()

    return redirect(url_for("main.equipamentos"))


@main.route("/equipamentos/<int:id>/excluir", methods=["POST"])
def excluir_equipamento(id):
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    recurso = Recurso.query.get_or_404(id)
    recurso.ativo = False
    db.session.commit()

    return redirect(url_for("main.equipamentos"))


@main.route("/usuarios/senha", methods=["POST"])
def alterar_senha_usuario():
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    usuario = Usuario.query.get_or_404(request.form["usuario_id"])
    nova_senha = request.form.get("senha")

    if nova_senha:
        usuario.senha = nova_senha
        db.session.commit()

    return redirect(url_for("main.home", tab="usuarios"))


@main.route("/usuarios/role", methods=["POST"])
def alterar_role_usuario():
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    usuario = Usuario.query.get_or_404(request.form["usuario_id"])
    nova_role = normalizar_texto(request.form.get("role"))
    roles_validas = ["usuario", "rh", "diretoria", "tecnico"]

    if nova_role in roles_validas:
        usuario.role = nova_role
        db.session.commit()

        if usuario.id == session.get("usuario_id"):
            session["role"] = nova_role

    return redirect(url_for("main.home", tab="usuarios"))


@main.route("/setores/adicionar", methods=["POST"])
def adicionar_setor():
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    nome = (request.form.get("nome") or "").strip()

    if nome:
        setor_existente = Setor.query.filter(db.func.lower(Setor.nome) == nome.lower()).first()

        if setor_existente:
            setor_existente.ativo = True
        else:
            db.session.add(Setor(nome=nome, ativo=True))

        db.session.commit()

    return redirect(url_for("main.home", tab="usuarios"))


@main.route("/setores/<int:id>/excluir", methods=["POST"])
def excluir_setor(id):
    bloqueio = exigir_tecnico()
    if bloqueio:
        return bloqueio

    setor = Setor.query.get_or_404(id)
    setor.ativo = False
    db.session.commit()

    return redirect(url_for("main.home", tab="usuarios"))


@main.route("/reserva")
def reserva():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    recurso_id = request.args.get("recurso_id", type=int)
    recursos = consulta_recursos_disponiveis().all()
    recurso_selecionado = Recurso.query.get(recurso_id) if recurso_id else None
    setores = consulta_setores_ativos().all()

    if recurso_selecionado and not recurso_disponivel_para_reserva(recurso_selecionado):
        flash("Este equipamento esta em manutencao ou indisponivel e nao pode ser reservado.", "erro")
        return redirect(url_for("main.reserva"))

    agenda_ocupada = montar_agenda_ocupada(recurso_id)
    pode_viagem = recurso_eh_veiculo(recurso_selecionado)

    return render_template(
        "reserva.html",
        recursos=recursos,
        recurso_selecionado=recurso_selecionado,
        agenda_ocupada=agenda_ocupada,
        pode_viagem=pode_viagem,
        setores=setores,
    )
