from io import BytesIO

from flask import redirect, render_template, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from controllers import main
from model import Reserva
from services.auth import exigir_historico_geral, pode_ver_relatorios
from services.pdf import gerar_pdf, gerar_pdf_historico
from services.relatorios import (
    consulta_reservas_filtrada,
    montar_dados_relatorio,
    montar_stats,
)
from services.reservas import ordenar_historico


def estilos_excel():
    return {
        "header_fill": PatternFill("solid", fgColor="00995C"),
        "header_font": Font(color="FFFFFF", bold=True),
        "body_font": Font(color="1F2D25"),
        "thin_border": Border(
            left=Side(style="thin", color="BFD8C8"),
            right=Side(style="thin", color="BFD8C8"),
            top=Side(style="thin", color="BFD8C8"),
            bottom=Side(style="thin", color="BFD8C8"),
        ),
    }


def formatar_planilha(ws, larguras=None):
    estilos = estilos_excel()
    max_row = ws.max_row or 1
    max_column = ws.max_column or 1

    for cell in ws[1]:
        cell.fill = estilos["header_fill"]
        cell.font = estilos["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = estilos["thin_border"]

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.border = estilos["thin_border"]
            cell.font = estilos["body_font"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index in range(1, max_row + 1):
        ws.row_dimensions[row_index].height = 24 if row_index == 1 else 36

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True

    for column_index in range(1, max_column + 1):
        column_letter = get_column_letter(column_index)
        if larguras and column_index in larguras:
            ws.column_dimensions[column_letter].width = larguras[column_index]
            continue

        column_cells = list(ws.iter_cols(min_col=column_index, max_col=column_index, max_row=max_row))[0]
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 14), 42)


def finalizar_excel(wb):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True
        ws.sheet_format.defaultRowHeight = 24


@main.route("/relatorios")
def relatorios():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not pode_ver_relatorios():
        return "Acesso negado", 403

    consulta, filtros = consulta_reservas_filtrada()
    registros = consulta.all()
    stats = montar_stats(registros)
    dados = montar_dados_relatorio(registros)

    return render_template(
        "relatorios.html",
        filtros=filtros,
        stats=stats,
        dados_setor=dados["setor"],
        dados_periodo=dados["periodo"],
        dados_recurso=dados["recurso"],
        dados_status=dados["status"],
        dados_hora=dados["hora"],
        dados_responsavel=dados["responsavel"],
        rankings=dados["rankings"],
    )


@main.route("/exportar/excel")
def exportar_excel():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not pode_ver_relatorios():
        return "Acesso negado", 403

    consulta, _ = consulta_reservas_filtrada()
    registros = consulta.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Base de dados"

    ws.append(["Data inicial", "Data devolucao", "Recurso", "Responsavel", "Setor", "Local/Motivo", "Status", "Hora inicial", "Hora final", "Observacao"])

    for registro in registros:
        ws.append(
            [
                registro.data_reserva.strftime("%d/%m/%Y"),
                (registro.data_volta or registro.data_reserva).strftime("%d/%m/%Y"),
                registro.recurso.nome if registro.recurso else "",
                registro.responsavel,
                registro.setor,
                registro.motivo,
                registro.status_label,
                registro.hora_inicio.strftime("%H:%M"),
                registro.hora_fim.strftime("%H:%M") if registro.hora_fim else "Viagem",
                registro.observacao,
            ]
        )

    formatar_planilha(
        ws,
        {
            1: 14,
            2: 16,
            3: 22,
            4: 24,
            5: 20,
            6: 30,
            7: 16,
            8: 14,
            9: 14,
            10: 36,
        },
    )

    resumo = wb.create_sheet("Resumo")
    stats = montar_stats(registros)
    dados = montar_dados_relatorio(registros)
    resumo.append(["Indicador", "Valor"])
    resumo.append(["Total de registros", stats["total"]])
    resumo.append(["Dias com uso", stats["diasComUso"]])
    resumo.append(["Media diaria", stats["mediaDiaria"]])
    resumo.append(["Recurso mais usado", stats["recursoTop"]])
    resumo.append(["Setor mais ativo", stats["setorTop"]])
    resumo.append(["Responsavel mais frequente", stats["requerenteTop"]])
    resumo.append(["Taxa de devolucao", f"{stats['taxaDevolucao']}%"])
    resumo.append(["Pendentes", stats["pendentes"]])
    resumo.append(["Em uso", stats["emUso"]])
    resumo.append(["Atrasados", stats["atrasados"]])
    resumo.append(["Devolvidos", stats["devolvidos"]])
    resumo.append(["Viagens", stats["viagens"]])

    start_row = resumo.max_row + 2
    resumo.cell(start_row, 1, "Distribuicao por status")
    resumo.cell(start_row, 1).font = Font(bold=True, color="00995C")
    for index, (label, valor) in enumerate(zip(dados["status"]["labels"], dados["status"]["valores"]), start=start_row + 1):
        resumo.cell(index, 1, label)
        resumo.cell(index, 2, valor)

    formatar_planilha(resumo, {1: 34, 2: 20})
    resumo.column_dimensions["A"].width = 30
    resumo.column_dimensions["B"].width = 18

    rankings = wb.create_sheet("Rankings")
    linha = 1
    estilos = estilos_excel()
    for titulo, itens in [
        ("Top recursos", dados["rankings"]["recursos"]),
        ("Top setores", dados["rankings"]["setores"]),
        ("Top responsaveis", dados["rankings"]["responsaveis"]),
    ]:
        rankings.cell(linha, 1, titulo)
        rankings.cell(linha, 1).font = Font(bold=True, color="00995C", size=13)
        linha += 1
        rankings.append(["Nome", "Reservas"])
        for cell in rankings[linha]:
            cell.fill = estilos["header_fill"]
            cell.font = estilos["header_font"]
            cell.alignment = Alignment(horizontal="center")
            cell.border = estilos["thin_border"]
        linha += 1
        for nome, valor in itens:
            rankings.cell(linha, 1, nome)
            rankings.cell(linha, 2, valor)
            rankings.cell(linha, 1).border = estilos["thin_border"]
            rankings.cell(linha, 2).border = estilos["thin_border"]
            rankings.cell(linha, 1).alignment = Alignment(vertical="top", wrap_text=True)
            rankings.cell(linha, 2).alignment = Alignment(horizontal="center", vertical="top")
            linha += 1
        linha += 2

    formatar_planilha(rankings, {1: 34, 2: 14})
    rankings.sheet_view.showGridLines = True

    finalizar_excel(wb)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main.route("/exportar/pdf")
def exportar_pdf():
    if "usuario" not in session:
        return redirect(url_for("main.login_page"))

    if not pode_ver_relatorios():
        return "Acesso negado", 403

    consulta, filtros = consulta_reservas_filtrada()
    registros = consulta.all()
    stats = montar_stats(registros)
    dados = montar_dados_relatorio(registros)

    return gerar_pdf(registros, filtros, stats, dados)


@main.route("/historico/exportar/excel")
def exportar_historico_excel():
    bloqueio = exigir_historico_geral()
    if bloqueio:
        return bloqueio

    registros = ordenar_historico(Reserva.query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historico"

    ws.append(["Data inicial", "Data devolucao", "Hora", "Responsavel", "Recurso", "Setor", "Local", "Status", "Observacao"])

    for registro in registros:
        hora_fim = registro.hora_fim.strftime("%H:%M") if registro.hora_fim else ("Viagem" if registro.viagem else "-")
        ws.append(
            [
                registro.data_reserva.strftime("%d/%m/%Y"),
                (registro.data_volta or registro.data_reserva).strftime("%d/%m/%Y"),
                f"{registro.hora_inicio.strftime('%H:%M')} - {hora_fim}",
                registro.responsavel or (registro.usuario.usuario if registro.usuario else "-"),
                registro.recurso.nome if registro.recurso else "-",
                registro.setor or "-",
                registro.motivo or "-",
                registro.status_label,
                registro.observacao or "-",
            ]
        )

    formatar_planilha(
        ws,
        {
            1: 14,
            2: 16,
            3: 20,
            4: 24,
            5: 22,
            6: 20,
            7: 30,
            8: 16,
            9: 36,
        },
    )

    finalizar_excel(wb)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="historico_geral_reservas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main.route("/historico/exportar/pdf")
def exportar_historico_pdf():
    bloqueio = exigir_historico_geral()
    if bloqueio:
        return bloqueio

    registros = ordenar_historico(Reserva.query).all()
    return gerar_pdf_historico(registros)
