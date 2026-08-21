# Permite criar o arquivo temporariamente na memória
from io import BytesIO

# Permite criar a planilha Excel
from openpyxl import Workbook

# Permite estilizar as células
from openpyxl.styles import Alignment, Font, PatternFill

# Busca os dados dos relatórios
from services.report_services import ReportService


# Gera os arquivos exportados dos relatórios
class ReportExportService:
    
    #crai o relatorio em excel
    @staticmethod
    def generate_excel(filters=None):
        #busca indicadores usando os filtros
        stats = ReportService.get_summary(filters)
        
        rankings = ReportService.get_rankings(filters)
        
        reservations = (
            ReportService.get_reservation_details(filters)
        )
        
        #cria uma nova planilha
        workbook = Workbook()
        
        #seleciona a primeira pagina da planilha
        sheet = workbook.active
        
        #define o nome da pag
        sheet.title = "Relatório"
        
        #cria o titulo principal
        sheet.merge_cells("A1:B1")
        sheet["A1"] = "Relatório de Reservas"
        
        #estiliza o titulo
        sheet["A1"].font = Font(
            bold=True,
            color="FFFFFF",
            size=16,
        )
        
        sheet["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor="00995C",
        )
        
        sheet["A1"].alignment = Alignment(
            horizontal="center",
        )
        
        #define os indicadores que seram exportados
        indicadores = [
            ("Total de reservas", stats["total"]),
            ("Pendentes", stats["pendentes"]),
            ("Atrasados", stats["atrasados"]),
            ("Taxa de devolução", f'{stats["taxaDevolucao"]}%'),
            ("Recurso mais utilizado", stats["recursoTop"]),
            ("Setor mais ativo", stats["setorTop"]),
            ("Média diária", stats["mediaDiaria"]),
            ("Dias com uso", stats["diasComUso"]),
            ("Em uso", stats["emUso"]),
            ("Reservados", stats["reservados"]),
            ("Viagens", stats["viagens"]),
            ("Devolvidos", stats["devolvidos"]),
        ]
        
        #adiciona os indicadores na planilha
        for row_number, indicador in enumerate(
            indicadores,
            start=3,
        ):
            sheet.cell(
                row=row_number,
                column=1,
                value=indicador[0]
            )
            
            sheet.cell(
                row=row_number,
                column=2,
                value=indicador[1]
            )
            
        #ajusta a largura da coluna 
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 24
        
        #cria pag de rankings
        rankings_sheet = workbook.create_sheet("Rankings")
        
        #cria os cabecalhos
        rankings_sheet["A1"] = "Recursos"
        rankings_sheet["B1"] = "Reservas"
        rankings_sheet["D1"] = "Setores"
        rankings_sheet["E1"] = "Reservas"
        rankings_sheet["G1"] = "Responsáveis"
        rankings_sheet["H1"] = "Reservas"
        
        #define os cabecalhos que seram estilizados
        header_cells = [
            "A1",
            "B1",
            "D1",
            "E1",
            "G1",
            "H1",
        ]
        
        #estiliza os cabecalhos
        for cell_name in header_cells:
            rankings_sheet[cell_name].font = Font(
                bold=True,
                color="FFFFFF"
            )
            
            rankings_sheet[cell_name].fill = PatternFill(
                fill_type="solid",
                fgColor="00995C"
            )
            
        #adiciona o ranking de recursos
        for row_number, item in enumerate(
            rankings["recursos"],
            start=2,
        ):
            rankings_sheet.cell(
                row=row_number,
                column=1,
                value=item[0]
            )
            
            rankings_sheet.cell(
                row=row_number,
                column=2,
                value=item[1]
            )
            
        #adiciona o ranking de setores
        for row_number, item in enumerate(
            rankings["setores"],
            start=2
        ):
            rankings_sheet.cell(
                row=row_number,
                column=4,
                value=item[0]
            )
            
            rankings_sheet.cell(
                row=row_number,
                column=5,
                value=item[1]
            )
            
        #adiciona o ranking de responsaveis
        for row_number, item in enumerate(
            rankings["responsaveis"],
            start=2
        ):
            
            rankings_sheet.cell(
                row=row_number,
                column=7,
                value=item[0]
            )
            
            rankings_sheet.cell(
                row=row_number,
                column=8,
                value=item[1]
            )
            
        #ajusta a largura da coluna
        for column_name in ["A", "D", "G"]:
            rankings_sheet.column_dimensions[
                column_name
            ].width = 28
            
        for column_name in ["B", "E", "H"]:
            rankings_sheet.column_dimensions[
                column_name
            ].width = 14
            
            
        #cria pag das reservas
        reservations_sheet = workbook.create_sheet("Reservas")
        
        #define as colunas da pag
        columns = [
            ("ID", "id"),
            ("Data", "data_reserva"),
            ("Data de volta", "data_volta"),
            ("Início", "hora_inicio"),
            ("Fim", "hora_fim"),
            ("Recurso", "recurso"),
            ("Responsável", "responsavel"),
            ("Setor", "setor"),
            ("Motivo", "motivo"),
            ("Observação", "observacao"),
            ("Viagem", "viagem"),
            ("Status", "status"),
        ]
        
        #cria os cabecalhos
        for column_number, column in enumerate(
            columns,
            start=1
        ):
            
            cell = reservations_sheet.cell(
                row=1,
                column=column_number,
                value=column[0]
            )
            
            #estiliza o cabecalho
            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )
            
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="00995C"
            )
            
            cell.alignment = Alignment(
                horizontal="center"
            )
            
        #adiciona as reservas
        for row_number, reservation in enumerate(
            reservations,
            start=2,
        ):
            for column_number, column in enumerate(
                columns,
                start=1,
            ):
                reservations_sheet.cell(
                    row=row_number,
                    column=column_number,
                    value=reservation.get(
                        column[1],
                        "",
                    ),
                )
            
        #mantem o cabecalho visivel
        reservations_sheet.freeze_panes = "A2"
        
        #ativa o filtro das colunas
        reservations_sheet.auto_filter.ref = (
            reservations_sheet.dimensions
        )
        
        #ajusta as larguras das colunas
        column_widths = {
            "A": 10,
            "B": 14,
            "C": 16,
            "D": 12,
            "E": 12,
            "F": 24,
            "G": 24,
            "H": 20,
            "I": 30,
            "J": 40,
            "K": 12,
            "L": 16,
        }
        
        #aplica as larguras
        for column_name, width in column_widths.items():
            reservations_sheet.column_dimensions[
                column_name
            ].width = width
            
        
        #cria o arquivo na memoria 
        excel_file = BytesIO()
        
        #Salva a planilha na memoria
        workbook.save(excel_file)
        
        #retorna ao inicio do arquivo
        excel_file.seek(0)
        
        return excel_file
